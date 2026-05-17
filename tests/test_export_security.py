"""Tests for csv_guard.safe_spreadsheet_cell, ExportService, and export query."""

from __future__ import annotations

import os
import sqlite3
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import openpyxl
import pytest

from taxops.core.paths import resolve_paths
from taxops.db.connection import open_connection
from taxops.db.migrate import apply_migrations
from taxops.repositories.audit_logs import AuditLogRepository
from taxops.repositories.document_requests import DocumentRequestsRepository
from taxops.security.csv_guard import safe_spreadsheet_cell
from taxops.services.audit import AuditService
from taxops.services.export import ExportService, ExportValidationError
from taxops.ui.action_registry import PAGE_DOC_REQUESTS, actions_for_page


# ── fixtures ───────────────────────────────────────────────────────────────────


@pytest.fixture
def conn(tmp_path):
    paths = resolve_paths(override_root=tmp_path / "data")
    paths.data_root.mkdir(parents=True, exist_ok=True)
    c = open_connection(paths.db_path)
    apply_migrations(c)
    yield c
    c.close()


@pytest.fixture
def repo(conn):
    return DocumentRequestsRepository(conn)


@pytest.fixture
def audit(conn):
    return AuditService(AuditLogRepository(conn), actor="test")


@pytest.fixture
def svc(repo, audit):
    return ExportService(repo=repo, audit=audit)


def _seed(conn) -> tuple[int, int, int]:
    """Create client, engagement, doc_request with 3 items; return (client_id, eng_id, req_id)."""
    conn.execute(
        "INSERT INTO clients (client_code, client_name, created_at, updated_at)"
        " VALUES ('C001', '測試客戶', '2026-01-01T00:00:00', '2026-01-01T00:00:00')"
    )
    client_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.execute(
        "INSERT INTO engagements (client_id, engagement_name, tax_type, period_name,"
        " status, created_at, updated_at)"
        " VALUES (?, '測試案件', 'vat', '202501', 'draft', '2026-01-01T00:00:00', '2026-01-01T00:00:00')",
        (client_id,),
    )
    eng_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.execute(
        "INSERT INTO document_requests (engagement_id, tax_type, period_name, status,"
        " follow_up_count, created_at, updated_at)"
        " VALUES (?, 'vat', '202501', 'not_requested', 0,"
        " '2026-01-01T00:00:00', '2026-01-01T00:00:00')",
        (eng_id,),
    )
    req_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    # 2 missing items, 1 accepted item
    for name, status in [("申報書", "missing"), ("發票", "incomplete"), ("收據", "accepted")]:
        conn.execute(
            "INSERT INTO document_request_items (request_id, item_name, item_status,"
            " created_at, updated_at)"
            " VALUES (?, ?, ?, '2026-01-01T00:00:00', '2026-01-01T00:00:00')",
            (req_id, name, status),
        )
    conn.commit()
    return client_id, eng_id, req_id


# ── safe_spreadsheet_cell ──────────────────────────────────────────────────────


@pytest.mark.parametrize("value,expected", [
    ("=SUM(A1:A10)", "'=SUM(A1:A10)"),
    ("+1234", "'+1234"),
    ("-1234", "'-1234"),
    ("@user", "'@user"),
    ("normal text", "normal text"),
    ("", ""),
    ("123", "123"),
    ("  =formula", "'  =formula"),
    ("\t+SUM(A1:A2)", "'\t+SUM(A1:A2)"),
])
def test_safe_spreadsheet_cell(value, expected):
    assert safe_spreadsheet_cell(value) == expected


def test_safe_spreadsheet_cell_nested_formula():
    v = "=HYPERLINK(\"http://evil.com\",\"click\")"
    assert safe_spreadsheet_cell(v).startswith("'")


# ── list_missing_items_for_export ──────────────────────────────────────────────


def test_export_query_excludes_accepted_items(conn, repo):
    _seed(conn)
    rows = repo.list_missing_items_for_export()
    statuses = {r["item_status"] for r in rows}
    assert "accepted" not in statuses


def test_export_query_includes_missing_and_incomplete(conn, repo):
    _seed(conn)
    rows = repo.list_missing_items_for_export()
    statuses = {r["item_status"] for r in rows}
    assert "missing" in statuses
    assert "incomplete" in statuses


def test_export_query_has_required_fields(conn, repo):
    _seed(conn)
    rows = repo.list_missing_items_for_export()
    assert len(rows) > 0
    row = rows[0]
    for field in (
        "client_code", "client_name", "tax_id", "engagement_name",
        "tax_type", "period_name", "item_name", "item_status",
        "owner", "due_date", "requested_at", "follow_up_count", "notes",
    ):
        assert field in row, f"Missing field: {field}"


def test_export_query_engagement_filter(conn, repo):
    _, eng_id, _ = _seed(conn)
    # add second client+engagement with its own item
    conn.execute(
        "INSERT INTO clients (client_code, client_name, created_at, updated_at)"
        " VALUES ('C002', '另一客戶', '2026-01-01T00:00:00', '2026-01-01T00:00:00')"
    )
    c2 = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.execute(
        "INSERT INTO engagements (client_id, engagement_name, tax_type, period_name,"
        " status, created_at, updated_at)"
        " VALUES (?, '另一案件', 'vat', '202501', 'draft', '2026-01-01T00:00:00', '2026-01-01T00:00:00')",
        (c2,),
    )
    eng2 = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.execute(
        "INSERT INTO document_requests (engagement_id, tax_type, period_name, status,"
        " follow_up_count, created_at, updated_at)"
        " VALUES (?, 'vat', '202501', 'not_requested', 0,"
        " '2026-01-01T00:00:00', '2026-01-01T00:00:00')",
        (eng2,),
    )
    req2 = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.execute(
        "INSERT INTO document_request_items (request_id, item_name, item_status,"
        " created_at, updated_at)"
        " VALUES (?, '其他文件', 'missing', '2026-01-01T00:00:00', '2026-01-01T00:00:00')",
        (req2,),
    )
    conn.commit()

    rows_all = repo.list_missing_items_for_export()
    rows_eng1 = repo.list_missing_items_for_export(engagement_id=eng_id)

    assert len(rows_all) > len(rows_eng1)
    eng_names = {r["engagement_name"] for r in rows_eng1}
    assert eng_names == {"測試案件"}


# ── ExportService ──────────────────────────────────────────────────────────────


def test_export_creates_xlsx_file(conn, svc, tmp_path):
    _seed(conn)
    out = tmp_path / "out.xlsx"
    count = svc.export_missing_items_xlsx(out)
    assert out.exists()
    assert count == 2  # 2 missing/incomplete, 1 accepted excluded


def test_export_xlsx_has_header_row(conn, svc, tmp_path):
    _seed(conn)
    out = tmp_path / "out.xlsx"
    svc.export_missing_items_xlsx(out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert "客戶代號" in header
    assert "缺件項目" in header
    assert "狀態" in header


def test_export_xlsx_data_rows_contain_client_info(conn, svc, tmp_path):
    _seed(conn)
    out = tmp_path / "out.xlsx"
    svc.export_missing_items_xlsx(out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    # row 2 is first data row; column 1 = client_code
    codes = [ws.cell(2, 1).value, ws.cell(3, 1).value]
    assert all(c == "C001" for c in codes)


def test_export_xlsx_formula_injection_escaped(conn, svc, tmp_path):
    """Item names starting with = must be escaped in the XLSX output."""
    conn.execute(
        "INSERT INTO clients (client_code, client_name, created_at, updated_at)"
        " VALUES ('C999', '注入測試', '2026-01-01T00:00:00', '2026-01-01T00:00:00')"
    )
    cid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.execute(
        "INSERT INTO engagements (client_id, engagement_name, tax_type, period_name,"
        " status, created_at, updated_at)"
        " VALUES (?, '注入案件', 'vat', '202501', 'draft', '2026-01-01T00:00:00', '2026-01-01T00:00:00')",
        (cid,),
    )
    eid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.execute(
        "INSERT INTO document_requests (engagement_id, tax_type, period_name, status,"
        " follow_up_count, created_at, updated_at)"
        " VALUES (?, 'vat', '202501', 'not_requested', 0,"
        " '2026-01-01T00:00:00', '2026-01-01T00:00:00')",
        (eid,),
    )
    rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.execute(
        "INSERT INTO document_request_items (request_id, item_name, item_status,"
        " created_at, updated_at)"
        " VALUES (?, '=HYPERLINK(\"http://evil.com\",\"click\")', 'missing',"
        " '2026-01-01T00:00:00', '2026-01-01T00:00:00')",
        (rid,),
    )
    conn.commit()

    out = tmp_path / "inject.xlsx"
    svc.export_missing_items_xlsx(out, engagement_id=eid)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    item_col = 7  # column G = item_name (1-indexed)
    item_value = ws.cell(2, item_col).value
    assert item_value is not None
    assert item_value.startswith("'"), f"Expected escaped formula, got: {item_value!r}"


def test_export_records_audit(conn, svc, tmp_path):
    _seed(conn)
    out = tmp_path / "audit_test.xlsx"
    svc.export_missing_items_xlsx(out)
    log = conn.execute(
        "SELECT * FROM audit_logs WHERE action = 'export.missing_items'"
    ).fetchone()
    assert log is not None


def test_export_empty_result_ok(conn, svc, tmp_path):
    """Export with no matching rows still produces a valid file with just headers."""
    out = tmp_path / "empty.xlsx"
    count = svc.export_missing_items_xlsx(out)
    assert count == 0
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.max_row == 1  # header only


# ── action registry contracts ──────────────────────────────────────────────────


def test_export_contract_in_registry():
    contracts = actions_for_page(PAGE_DOC_REQUESTS)
    labels = {c.button_label for c in contracts}
    assert "匯出缺件清單" in labels


def test_export_contract_has_audit_action():
    for c in actions_for_page(PAGE_DOC_REQUESTS):
        if c.button_label == "匯出缺件清單":
            assert c.audit_action == "export.missing_items"
            assert c.service is not None
            assert c.repository is not None
            break


# ── UI handler integration ────────────────────────────────────────────────────


@pytest.fixture(scope="session")
def qapp():
    from PySide6.QtWidgets import QApplication

    return QApplication.instance() or QApplication([])


def _make_container(tmp_path):
    from taxops.services.container import build_container

    paths = resolve_paths(override_root=tmp_path / "ui_data")
    paths.data_root.mkdir(parents=True, exist_ok=True)
    paths.attachments_dir.mkdir(parents=True, exist_ok=True)
    c = open_connection(paths.db_path)
    apply_migrations(c)
    return build_container(paths, c)


def test_document_requests_page_export_handler_writes_file_and_audit(
    qapp, tmp_path, monkeypatch
):
    from PySide6.QtWidgets import QFileDialog, QMessageBox
    from taxops.ui.pages.document_requests_page import DocumentRequestsPage

    container = _make_container(tmp_path)
    try:
        _, eng_id, _ = _seed(container.conn)
        out = tmp_path / "ui-export.xlsx"
        messages: list[tuple[str, str]] = []

        monkeypatch.setattr(
            QFileDialog,
            "getSaveFileName",
            lambda *args, **kwargs: (str(out), "Excel 檔案 (*.xlsx)"),
        )
        monkeypatch.setattr(
            QMessageBox,
            "information",
            lambda _parent, title, text: messages.append((title, text)),
        )

        page = DocumentRequestsPage(container)
        page.load_engagement(eng_id)
        assert page._export_btn.isEnabled()

        page._on_export()

        assert out.exists()
        wb = openpyxl.load_workbook(out)
        assert wb.active.max_row == 3
        assert messages and messages[-1][0] == "匯出完成"
        log = container.conn.execute(
            "SELECT * FROM audit_logs WHERE action = 'export.missing_items'"
        ).fetchone()
        assert log is not None
    finally:
        container.close()


def test_document_requests_page_export_handler_cancel_does_not_write(
    qapp, tmp_path, monkeypatch
):
    from PySide6.QtWidgets import QFileDialog
    from taxops.ui.pages.document_requests_page import DocumentRequestsPage

    container = _make_container(tmp_path)
    try:
        _, eng_id, _ = _seed(container.conn)
        monkeypatch.setattr(
            QFileDialog,
            "getSaveFileName",
            lambda *args, **kwargs: ("", ""),
        )

        page = DocumentRequestsPage(container)
        page.load_engagement(eng_id)
        page._on_export()

        log = container.conn.execute(
            "SELECT * FROM audit_logs WHERE action = 'export.missing_items'"
        ).fetchone()
        assert log is None
    finally:
        container.close()
